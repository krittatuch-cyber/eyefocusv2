import urllib.request
import openpyxl
import os

url = "https://docs.google.com/spreadsheets/d/1yrwE6f3q2kSs1arvvfFLbyc0sDbfUMgP9C4G8ad7hSs/export?format=xlsx"
temp_file = "temp_sheets.xlsx"

print("Downloading Excel...")
urllib.request.urlretrieve(url, temp_file)
print("Downloaded. Opening workbook...")

wb = openpyxl.load_workbook(temp_file)
print("Sheets found:", wb.sheetnames)

output_path = r"C:\Users\kritt\.gemini\antigravity\brain\c5c0654e-5f2d-45a0-a55b-e98338a0a4e6\sheets_content.md"
with open(output_path, "w", encoding="utf-8") as f:
    f.write("# Google Sheet Content\n\n")
    for sheet_name in wb.sheetnames:
        f.write(f"## Sheet: {sheet_name}\n\n")
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row_str = [str(cell).replace("\n", " ").replace("|", "\\|") if cell is not None else "" for cell in row]
            if any(row_str):
                f.write("| " + " | ".join(row_str) + " |\n")
        f.write("\n\n")

print("Done. Output written to:", output_path)
# Clean up temp file
if os.path.exists(temp_file):
    os.remove(temp_file)
